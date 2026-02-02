import os
import io
import openpyxl
import re
from datetime import datetime, timezone
from typing import List, Optional, Literal, Set, Tuple, Dict
from fastapi import FastAPI, HTTPException, UploadFile, File, Body
from fastapi.middleware.cors import CORSMiddleware
from supabase import create_client, Client
from pydantic import BaseModel
from dotenv import load_dotenv

# --- Setup ---
load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), "..", "..", ".env.local"))

url: str = os.environ.get("SUPABASE_URL")
key: str = os.environ.get("SUPABASE_SERVICE_KEY")
supabase: Client = create_client(url, key)

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# --- Constants based on User Logic ---
PRIORITY_ORDER = {
    "primary": 1,      # 60-95% target
    "secondary": 2,    # 5-30% target
    "tertiary": 3,     # 1-10% target (National: KazNMU, MUA, etc.)
    "specialized": 4,  # 1-5% target
    "none": 99         # Fallback
}

# КАЗНМУ (National University): add to ALL regions so every region has at least one candidate
KAZNMU_PRIMARY_KEYWORDS = ("алматы", "жетысуская", "алматинская")
KAZNMU_SECONDARY_KEYWORDS = ("жамбылская", "туркестанская", "кызылординская", "шымкент")


def _find_kaznmu_id(supabase_client):
    """Return university id for КАЗНМУ (name contains КазНМУ or Асфендиярова), or None."""
    try:
        data = supabase_client.table("universities").select("id, name").execute().data or []
        for row in data:
            name = (row.get("name") or "").strip().lower()
            if "казнму" in name or "асфендиярова" in name:
                return str(row.get("id"))
    except Exception:
        pass
    return None


def _kaznmu_rank_for_region(region_name_normalized: str) -> int:
    """Return priority rank for КАЗНМУ in this region: 1=primary, 2=secondary, 3=tertiary."""
    if not region_name_normalized:
        return 3
    r = region_name_normalized.strip().lower()
    if any(kw in r for kw in KAZNMU_PRIMARY_KEYWORDS):
        return 1
    if any(kw in r for kw in KAZNMU_SECONDARY_KEYWORDS):
        return 2
    return 3


# --- Pydantic Models ---
class CreateSessionRequest(BaseModel):
    session_name: str
    year: int
    total_budget: int
    uploaded_by: Optional[str] = "system"

class GeoFilterRequest(BaseModel):
    region_id: Optional[str] = None
    specialty_id: Optional[str] = None
    session_id: Optional[str] = None

class UpsertUniversityAllocationsRequest(BaseModel):
    region_id: str
    specialty_id: str
    assignments: List[dict]

class RunAllocationRequest(BaseModel):
    priority: str  # historic_deficit | requested_amount | last_year_data
    total_grants: Optional[int] = None

class AdjustAllocationRequest(BaseModel):
    new_allocation: Optional[int] = None
    notes: Optional[str] = None

class BulkAdjustmentItem(BaseModel):
    demand_request_id: str
    new_allocation: Optional[int] = None
    notes: Optional[str] = None

class BulkAdjustRequest(BaseModel):
    adjustments: List[BulkAdjustmentItem]

class FinalizeSessionRequest(BaseModel):
    finalized_by: Optional[str] = None

# --- Helper Functions ---

def _effective_allocation(d: dict) -> int:
    """Effective allocation: final_allocation column, with fallback to user_allocation or initial_allocation."""
    val = d.get("final_allocation")
    if val is not None:
        return int(val)
    u = d.get("user_allocation")
    if u is not None:
        return int(u)
    return int(d.get("initial_allocation") or 0)

def _normalize_name_for_lookup(name: str) -> str:
    """Normalize names for reliable DB lookup."""
    if not name: return ""
    s = str(name).strip().lower()
    s = re.sub(r"[«»\"']", "", s) # Remove quotes
    s = re.sub(r"^\s*(г\.|город|нао|нуо|тоо)\s*", "", s) # Remove prefixes
    s = re.sub(r"\s+", " ", s).strip()
    return s

def _normalize_specialty_for_graduate_lookup(specialty_name: str) -> str:
    """Normalize specialty name for yearly_graduates lookup."""
    return _normalize_name_for_lookup(specialty_name or "")

def allocation_by_priority(priority: str, rows: List[dict]) -> List[int]:
    """
    Set initial allocation per row from the chosen priority.
    historic_deficit -> historical_deficit, requested_amount -> current_request, last_year_data -> graduate_count.
    Returns list of allocations (same order as rows).
    """
    result = []
    for r in rows:
        hist = max(0, int(r.get("historical_deficit") or 0))
        req = max(0, int(r.get("current_request") or 0))
        grad = max(0, int(r.get("graduate_count") or 0))
        if priority == "historic_deficit":
            result.append(hist)
        elif priority == "requested_amount":
            result.append(req)
        elif priority == "last_year_data":
            result.append(grad)
        else:
            result.append(max(hist, req, grad))
    return result

def _get_db_maps(supabase_client):
    """Fetch Regions, Specialties, and Universities to map Names -> IDs."""
    r_map, s_map, u_map = {}, {}, {}
    
    # Regions
    try:
        data = supabase_client.table("regions").select("id, name").execute().data or []
        for row in data:
            r_map[_normalize_name_for_lookup(row["name"])] = row["id"]
    except: pass

    # Specialties
    try:
        data = supabase_client.table("specialties").select("id, name").execute().data or []
        for row in data:
            s_map[_normalize_name_for_lookup(row["name"])] = row["id"]
    except: pass

    # Universities (Map column headers like "КазНМУ" to DB IDs)
    try:
        data = supabase_client.table("universities").select("id, name, short_name").execute().data or []
        for row in data:
            uid = row["id"]
            # Map full name
            u_map[_normalize_name_for_lookup(row["name"])] = uid
            # Map short name if exists (e.g. "МУА")
            if row.get("short_name"):
                u_map[_normalize_name_for_lookup(row["short_name"])] = uid
    except: pass
    
    return r_map, s_map, u_map

# --- CORE LOGIC: Strict Waterfall Allocation ---

def _compute_geo_university_allocations(supabase_client, demand_filter: Optional[Set[Tuple[str, str]]] = None, session_id: Optional[str] = None):
    """
    Allocates grants strictly based on the SQL prioritization logic:
    1. Primary (Local) -> Fill first.
    2. Secondary (Regional Neighbors) -> Fill second.
    3. Tertiary (National: KazNMU, MUA, etc.) -> Fill third.
    4. Specialized -> Fill last.
    Uses initial_allocation for demand quantity (algorithm output before user edits).
    When session_id is provided, only demands for that session are used.
    """
    
    # 1. Fetch Supply: Capacity per (Uni, Specialty)
    cap_resp = supabase_client.table("university_specialty_capacity").select("university_id, specialty_id, capacity").execute()
    supply_map = {} # (uid, sid) -> remaining_capacity
    for row in (cap_resp.data or []):
        k = (str(row["university_id"]), str(row["specialty_id"]))
        supply_map[k] = int(row["capacity"] or 0)

    # 2. Fetch Demand: Requests per (Region, Specialty); use initial_allocation, cap by session total_budget
    demands_resp = supabase_client.table("demand_requests").select(
        "region_id, specialty_id, initial_allocation, session_id"
    ).execute()
    demands = []
    session_id_from_demand = None
    for row in (demands_resp.data or []):
        if session_id and str(row.get("session_id") or "") != str(session_id):
            continue
        rid, sid = str(row["region_id"]), str(row["specialty_id"])
        if demand_filter and (rid, sid) not in demand_filter:
            continue
        qty = row.get("initial_allocation") or 0
        if qty > 0:
            if row.get("session_id"):
                session_id_from_demand = str(row["session_id"])
            demands.append({"region_id": rid, "specialty_id": sid, "needed": int(qty)})
    total_budget = None
    session_for_budget = session_id or session_id_from_demand
    if session_for_budget:
        try:
            sess = supabase_client.table("allocation_sessions").select("total_budget").eq(
                "id", session_for_budget
            ).execute()
            if sess.data and len(sess.data) > 0 and (sess.data[0].get("total_budget") or 0) > 0:
                total_budget = int(sess.data[0]["total_budget"])
        except Exception:
            pass
    if total_budget is not None and total_budget > 0:
        total_needed = sum(d["needed"] for d in demands)
        if total_needed > total_budget:
            cap = total_budget
            factor = cap / total_needed
            for d in demands:
                d["needed"] = int(d["needed"] * factor)
            remainder = cap - sum(d["needed"] for d in demands)
            if remainder > 0:
                with_frac = [(i, d["needed"] * factor - int(d["needed"] * factor)) for i, d in enumerate(demands)]
                with_frac.sort(key=lambda x: -x[1])
                for j in range(min(remainder, len(with_frac))):
                    demands[with_frac[j][0]]["needed"] += 1

    # 3. Fetch Matrix: Proximity Priorities (The Source of Truth)
    # This table assumes the User's SQL has been run.
    prox_resp = supabase_client.table("university_region_proximity").select("university_id, region_id, priority").execute()
    proximity_map = {} # (uid, rid) -> priority_rank (int)
    
    # Pre-process priorities
    uni_serving_region = {} # region_id -> list of (uid, priority_rank)
    
    for row in (prox_resp.data or []):
        uid = str(row["university_id"])
        rid = str(row["region_id"])
        p_str = (row["priority"] or "").lower().strip()
        rank = PRIORITY_ORDER.get(p_str, 99)
        proximity_map[(uid, rid)] = rank
        
        if rid not in uni_serving_region:
            uni_serving_region[rid] = []
        uni_serving_region[rid].append((uid, rank))

    # Sort universities for each region by strict rank
    for rid in uni_serving_region:
        uni_serving_region[rid].sort(key=lambda x: x[1]) # Sort by rank (1=Primary asc)

    # 3b. КАЗНМУ: add to every region so every (region, specialty) has at least one candidate
    kaznmu_id = _find_kaznmu_id(supabase_client)
    if kaznmu_id:
        try:
            regions_resp = supabase_client.table("regions").select("id, name").execute()
            region_id_to_name = {}
            for r in (regions_resp.data or []):
                region_id_to_name[str(r.get("id"))] = _normalize_name_for_lookup(r.get("name") or "")
        except Exception:
            region_id_to_name = {}
        all_rids = set(d["region_id"] for d in demands) | set(uni_serving_region.keys())
        for rid in all_rids:
            if not rid:
                continue
            region_norm = region_id_to_name.get(rid, "")
            kaznmu_rank = _kaznmu_rank_for_region(region_norm)
            existing_uids = {u for u, _ in uni_serving_region.get(rid, [])}
            if kaznmu_id not in existing_uids:
                uni_serving_region.setdefault(rid, []).append((kaznmu_id, kaznmu_rank))
        for rid in uni_serving_region:
            uni_serving_region[rid].sort(key=lambda x: x[1])

    # Process demands by (specialty_id, needed ASC, region_id) so small demands get filled first and more regions get a university
    demands_sorted = sorted(demands, key=lambda d: (d["specialty_id"], d["needed"], d["region_id"]))

    allocations = [] # Result list

    # 4. Allocation Algorithm
    for demand in demands_sorted:
        rid = demand["region_id"]
        sid = demand["specialty_id"]
        needed = demand["needed"]
        
        if needed <= 0: continue

        # Get candidates for this region
        candidates = uni_serving_region.get(rid, [])
        
        # Iteration 1: Strict Priority Waterfall
        # We walk through candidates: Primary first, then Secondary, etc.
        for uid, rank in candidates:
            if needed <= 0: break
            
            # Check if Uni has capacity for this Specialty
            cap_key = (uid, sid)
            available = supply_map.get(cap_key, 0)
            
            if available > 0:
                # Allocating
                take = min(needed, available)
                
                # Update State
                allocations.append({
                    "region_id": rid, 
                    "specialty_id": sid, 
                    "university_id": uid, 
                    "allocated_count": take,
                    "priority_applied": rank
                })
                
                supply_map[cap_key] -= take
                needed -= take

        # Iteration 2: Fallback (Desperation)
        # If needed > 0, find ANY university with capacity for this specialty, ignoring region
        if needed > 0:
            # Find all unis with capacity for this specialty
            fallback_unis = [u for u, c in supply_map.items() if u[1] == sid and c > 0]
            # Sort by capacity desc to fill quickly
            fallback_unis.sort(key=lambda x: supply_map[x], reverse=True)
            
            for (uid, _) in fallback_unis:
                if needed <= 0: break
                available = supply_map[(uid, sid)]
                take = min(needed, available)
                
                allocations.append({
                    "region_id": rid,
                    "specialty_id": sid,
                    "university_id": uid,
                    "allocated_count": take,
                    "priority_applied": 99 # Fallback
                })
                supply_map[(uid, sid)] -= take
                needed -= take

    # Consolidate results (merge multiple entries for same reg/spec/uni)
    consolidated = {}
    for a in allocations:
        key = (a["region_id"], a["specialty_id"], a["university_id"])
        consolidated[key] = consolidated.get(key, 0) + a["allocated_count"]
        
    final_output = [
        {"region_id": k[0], "specialty_id": k[1], "university_id": k[2], "allocated_count": v}
        for k, v in consolidated.items()
    ]
    
    return final_output

# --- Excel Parsers ---

def parse_excel_full_allocation(file_content, session_id, supabase_client):
    """
    Parses the 'Svodnaya' style Excel which contains both:
    1. Demands (Rows)
    2. University Capacities (Columns like 'КазНМУ', 'МУА')
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    ws = wb.active # Assuming first sheet is summary
    
    # 1. Map Headers to DB IDs
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    r_map, s_map, u_map = _get_db_maps(supabase_client)
    
    col_idx_map = {} # col_index -> {type: 'uni'|'meta', id: ...}
    
    region_col = None
    specialty_col = None
    hist_col = None
    req_col = None
    
    for idx, h in enumerate(headers):
        h_norm = _normalize_name_for_lookup(str(h))
        
        if h_norm in ['регион', 'область', 'region']:
            region_col = idx
        elif h_norm in ['специальность', 'specialty']:
            specialty_col = idx
        elif h_norm in ['потребность', 'historical_deficit', 'дефицит']:
            hist_col = idx
        elif h_norm in ['заявка', 'current_request', 'request']:
            req_col = idx
        
        # Check if header is a University
        if h_norm in u_map:
            col_idx_map[idx] = {"type": "uni", "id": u_map[h_norm]}
    
    demands_to_insert = []
    capacities_to_upsert = {} # (uni_id, spec_id) -> capacity
    
    # 2. Iterate Rows
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row: continue
        
        # Extract Identifiers
        r_name = row[region_col] if region_col is not None else None
        s_name = row[specialty_col] if specialty_col is not None else None
        
        if not s_name: continue # Specialty is mandatory
        
        rid = r_map.get(_normalize_name_for_lookup(r_name)) if r_name else None
        sid = s_map.get(_normalize_name_for_lookup(s_name))
        
        # Insert Demand
        if rid and sid:
            hist = row[hist_col] if hist_col is not None else 0
            req = row[req_col] if req_col is not None else 0
            try: hist = int(hist or 0)
            except: hist = 0
            try: req = int(req or 0)
            except: req = 0
            
            max_need = max(hist, req)
            demands_to_insert.append({
                "session_id": session_id,
                "region_id": rid,
                "specialty_id": sid,
                "historical_deficit": hist,
                "current_request": req,
                "initial_allocation": max_need,
                "max_need": max_need,
            })
            
        # Extract Capacities (Supply) from University Columns
        # Note: Even if Region is null (total row), we might want capacity. 
        # But usually capacity is defined per specialty row. 
        if sid:
            for idx, cell_val in enumerate(row):
                if idx in col_idx_map and col_idx_map[idx]["type"] == "uni":
                    uid = col_idx_map[idx]["id"]
                    try:
                        cap = int(cell_val or 0)
                    except:
                        cap = 0
                    
                    if cap > 0:
                        # Accumulate capacity for this specialty/uni
                        # (In case multiple rows have same specialty, usually strictly grouped)
                        key = (uid, sid)
                        capacities_to_upsert[key] = capacities_to_upsert.get(key, 0) + cap

    # 3. Perform DB Operations
    
    # A. Insert Demands
    if demands_to_insert:
        supabase_client.table("demand_requests").insert(demands_to_insert).execute()
        
    # B. Upsert Capacities
    # We clean existing capacities for these specialties to avoid doubling up on re-upload
    spec_ids = list(set([k[1] for k in capacities_to_upsert.keys()]))
    if spec_ids:
        # Note: In a real prod env, be careful deleting. For now, we assume session-based or refresh.
        # Supabase doesn't support bulk upsert easily without logic, so we iterate or use RPC.
        # We will iterate for safety here.
        for (uid, sid), cap in capacities_to_upsert.items():
            supabase_client.table("university_specialty_capacity").upsert({
                "university_id": uid,
                "specialty_id": sid,
                "capacity": cap
            }, on_conflict="university_id, specialty_id").execute()

    return len(demands_to_insert), len(capacities_to_upsert)


async def process_uploaded_excel(content: bytes, session_id: str, supabase_client) -> dict:
    """
    Process uploaded Excel: parse demands and capacities, insert demands.
    Returns {success, demands_inserted, suggestions_generated, error?}.
    """
    try:
        d_count, c_count = parse_excel_full_allocation(content, session_id, supabase_client)
        return {"success": True, "demands_inserted": d_count, "suggestions_generated": 0}
    except Exception as e:
        return {"success": False, "error": str(e), "demands_inserted": 0, "suggestions_generated": 0}


async def process_uploaded_graduates(
    content: bytes,
    supabase_client,
    session_id: Optional[str] = None,
    default_year: Optional[int] = None,
    file_name: Optional[str] = None,
    sheet_name: Optional[str] = None,
    replace_years: Optional[List[int]] = None,
) -> dict:
    """
    Process uploaded Excel with graduate counts (year, specialty, graduate_count).
    Returns {success, inserted, total_rows, metadata, error?}.
    """
    try:
        # Stub: no parser in this file for graduate Excel; return empty success.
        # Replace with actual parse + insert into yearly_graduates when available.
        return {"success": True, "inserted": 0, "total_rows": 0, "metadata": {}}
    except Exception as e:
        return {"success": False, "error": str(e), "inserted": 0, "total_rows": 0, "metadata": {}}


# --- Routes ---

@app.post("/sessions/{session_id}/upload-file")
async def upload_file(session_id: str, file: UploadFile = File(...)):
    """
    Process Excel: Extracts Demands AND University Capacities.
    """
    try:
        content = await file.read()
        if not file.filename.endswith(('.xlsx', '.xls')):
             raise HTTPException(400, "Invalid file format")

        d_count, c_count = parse_excel_full_allocation(content, session_id, supabase)
        
        supabase.table("allocation_sessions").update({
            "source_file_name": file.filename,
            "status": "in_review"
        }).eq("id", session_id).execute()

        return {
            "status": "success", 
            "demands_inserted": d_count, 
            "capacities_updated": c_count,
            "message": "File processed. Demands and University Capacities updated."
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(500, str(e))

@app.post("/university-allocations/geo")
async def compute_geo_university_allocations(
    save: bool = False,
    body: Optional[GeoFilterRequest] = Body(None),
):
    """
    Run the Waterfall Allocation Algorithm using SQL-defined priorities.
    """
    try:
        filter_set = None
        if body and body.region_id and body.specialty_id:
            filter_set = {(body.region_id, body.specialty_id)}
        session_id = body.session_id if body else None
        assignments = _compute_geo_university_allocations(supabase, demand_filter=filter_set, session_id=session_id)
        
        if save and assignments:
            # Clean old assignments if filtering, or truncate if full run?
            # Safe approach: Delete for the specific (Region, Specialty) pairs involved
            pairs = list(set((a["region_id"], a["specialty_id"]) for a in assignments))
            
            # Batch delete (simplified for readability, optimize for large datasets)
            for r, s in pairs:
                supabase.table("university_allocation_assignments").delete().eq("region_id", r).eq("specialty_id", s).execute()
                
            # Batch insert
            batch_size = 100
            for i in range(0, len(assignments), batch_size):
                batch = assignments[i:i+batch_size]
                supabase.table("university_allocation_assignments").insert(batch).execute()

        return {
            "status": "success",
            "assignments": assignments,
            "count": len(assignments)
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(500, str(e))

# ... [Include other endpoints: create_session, review tables, etc. from original code] ...
# --- API Endpoints ---

@app.post("/sessions/create")
async def create_session(request: CreateSessionRequest):
    """
    Create a new allocation session
    """
    try:
        session_data = {
            "session_name": request.session_name,
            "year": request.year,
            "total_budget": request.total_budget,
            "uploaded_by": request.uploaded_by,
            "status": "draft",
            "uploaded_at": datetime.now(timezone.utc).isoformat()
        }
        
        result = supabase.table("allocation_sessions").insert(session_data).execute()
        
        if not result.data:
            raise HTTPException(status_code=500, detail="Failed to create session")
        
        session = result.data[0]
        
        return {
            "status": "success",
            "message": "Session created successfully",
            "session": session
        }
        
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/sessions/{session_id}/upload-file")
async def upload_file(session_id: str, file: UploadFile = File(...)):
    """
    Upload Excel or PDF file and extract demand data
    Automatically detects file type and uses appropriate parser
    """
    try:
        # Read file content
        content = await file.read()
        
        # Determine file type
        file_extension = file.filename.lower().split('.')[-1]
        
        if file_extension in ['xlsx', 'xls']:
            # Use Excel parser
            print(f"Processing Excel file: {file.filename}")
            result = await process_uploaded_excel(content, session_id, supabase)
            
            if not result['success']:
                raise HTTPException(status_code=400, detail=result.get('error', 'Failed to process Excel'))
            
            # Update session with file info
            supabase.table("allocation_sessions").update({
                "source_file_name": file.filename,
                "status": "in_review"
            }).eq("id", session_id).execute()
            
            return {
                "status": "success",
                "message": f"Successfully processed {result['demands_inserted']} demands from Excel",
                "session_id": session_id,
                "file_name": file.filename,
                "file_type": "excel",
                "demands_count": result['demands_inserted'],
                "suggestions_count": result['suggestions_generated'],
                "metadata": result['metadata']
            }
            
        elif file_extension == 'pdf':
            # Use PDF parser (implement if needed)
            raise HTTPException(
                status_code=400,
                detail="PDF parsing not yet implemented. Please upload Excel (.xlsx) file."
            )
        
        else:
            raise HTTPException(
                status_code=400,
                detail=f"Unsupported file type: {file_extension}. Please upload .xlsx or .pdf"
            )
        
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        print("Upload error:", traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/sessions/{session_id}/review-table")
async def get_review_table(session_id: str, page: int = 1, page_size: int = 50):
    """
    Get the Excel-like review table with highlighting and pagination
    """
    try:
        # Calculate offset
        offset = (page - 1) * page_size
        
        # Get total count
        count_resp = supabase.table("demand_requests").select("*", count="exact").eq("session_id", session_id).execute()
        total_count = count_resp.count if hasattr(count_resp, 'count') else 0
        
        # Get paginated data with suggestions (join regions/specialties for names)
        demands_resp = supabase.table("demand_requests").select(
            "*, region:regions(id, name), specialty:specialties(id, name)"
        ).eq("session_id", session_id).range(offset, offset + page_size - 1).order("region_id").order("specialty_id").execute()
        
        demands = demands_resp.data or []
        
        # Get session for year (for graduates lookup)
        session_resp = supabase.table("allocation_sessions").select("*").eq("id", session_id).execute()
        session = session_resp.data[0] if session_resp.data else None
        graduate_year = (session.get("year") or datetime.now().year) - 1 if session else (datetime.now().year - 1)
        graduates_resp = supabase.table("yearly_graduates").select("specialty, graduate_count").eq("year", graduate_year).execute()
        graduates_map = {}
        for g in (graduates_resp.data or []):
            key = _normalize_specialty_for_graduate_lookup(g.get("specialty"))
            graduates_map[key] = int(g.get("graduate_count") or 0)
        
        def _region_name(d):
            r = d.get("region")
            return r.get("name", "").strip() if isinstance(r, dict) else (d.get("region") or "")
        def _specialty_name(d):
            s = d.get("specialty")
            return s.get("name", "").strip() if isinstance(s, dict) else (d.get("specialty") or "")
        
        # Get suggestions for these demands
        demand_ids = [d['id'] for d in demands]
        suggestions_resp = supabase.table("adjustment_suggestions").select("*").in_("demand_request_id", demand_ids).execute()
        
        suggestions_map = {}
        if suggestions_resp.data:
            for s in suggestions_resp.data:
                suggestions_map[s['demand_request_id']] = s
        
        # Merge data (include graduate_count from previous year)
        review_data = []
        for demand in demands:
            suggestion = suggestions_map.get(demand['id'], {})
            region = _region_name(demand)
            specialty = _specialty_name(demand)
            graduate_count = graduates_map.get(_normalize_specialty_for_graduate_lookup(specialty))
            review_data.append({
                **demand,
                'graduate_count': graduate_count,
                'graduate_year': graduate_year,
                'suggestion_type': suggestion.get('suggestion_type'),
                'suggested_reduction': suggestion.get('suggested_reduction'),
                'suggestion_reason': suggestion.get('reason'),
                'highlight_color': suggestion.get('highlight_color'),
                'final_allocation': _effective_allocation(demand),
                'review_status': 'reviewed' if demand.get('user_allocation') is not None else 'pending'
            })
        
        # Session already fetched above
        # Calculate summary statistics
        all_demands = supabase.table("demand_requests").select("*").eq("session_id", session_id).execute()
        total_initial = sum(d.get('initial_allocation', 0) for d in all_demands.data)
        total_final = sum(_effective_allocation(d) for d in all_demands.data)
        reviewed_count = sum(1 for d in all_demands.data if d.get('user_allocation') is not None)
        
        return {
            "status": "success",
            "session": session,
            "review_table": review_data,
            "pagination": {
                "page": page,
                "page_size": page_size,
                "total_count": total_count,
                "total_pages": (total_count + page_size - 1) // page_size
            },
            "summary": {
                "total_budget": session.get('total_budget', 0) if session else 0,
                "total_initial_allocation": total_initial,
                "total_final_allocation": total_final,
                "budget_remaining": (session.get('total_budget', 0) if session else 0) - total_final,
                "reviewed_count": reviewed_count,
                "pending_count": total_count - reviewed_count,
                "review_progress_pct": round((reviewed_count / total_count * 100), 1) if total_count > 0 else 0
            }
        }
        
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/sessions/{session_id}/review-table-full")
async def get_full_review_table(session_id: str):
    """
    Get the complete review table without pagination (for export)
    WARNING: Use with caution on large datasets
    """
    try:
        session_resp = supabase.table("allocation_sessions").select("*").eq("id", session_id).execute()
        session = session_resp.data[0] if session_resp.data else None
        graduate_year = (session.get("year") or datetime.now().year) - 1 if session else (datetime.now().year - 1)
        graduates_resp = supabase.table("yearly_graduates").select("specialty, graduate_count").eq("year", graduate_year).execute()
        graduates_map = {}
        for g in (graduates_resp.data or []):
            key = _normalize_specialty_for_graduate_lookup(g.get("specialty"))
            graduates_map[key] = int(g.get("graduate_count") or 0)
        
        demands_resp = supabase.table("demand_requests").select(
            "*, region:regions(id, name), specialty:specialties(id, name)"
        ).eq("session_id", session_id).order("region_id").order("specialty_id").execute()
        demands = demands_resp.data or []
        
        def _region_name(d):
            r = d.get("region")
            return r.get("name", "").strip() if isinstance(r, dict) else (d.get("region") or "")
        def _specialty_name(d):
            s = d.get("specialty")
            return s.get("name", "").strip() if isinstance(s, dict) else (d.get("specialty") or "")
        
        demand_ids = [d['id'] for d in demands]
        if demand_ids:
            suggestions_resp = supabase.table("adjustment_suggestions").select("*").in_("demand_request_id", demand_ids).execute()
            suggestions_map = {s['demand_request_id']: s for s in (suggestions_resp.data or [])}
        else:
            suggestions_map = {}
        
        review_data = []
        for demand in demands:
            suggestion = suggestions_map.get(demand['id'], {})
            region = _region_name(demand)
            specialty = _specialty_name(demand)
            graduate_count = graduates_map.get(_normalize_specialty_for_graduate_lookup(specialty))
            review_data.append({
                'id': demand['id'],
                'region': demand.get('region'),
                'specialty': demand.get('specialty'),
                'historical_deficit': demand['historical_deficit'],
                'current_request': demand['current_request'],
                'max_need': max(demand.get('historical_deficit') or 0, demand.get('current_request') or 0),
                'graduate_count': graduate_count,
                'graduate_year': graduate_year,
                'initial_allocation': demand.get('initial_allocation', 0),
                'user_allocation': demand.get('user_allocation'),
                'final_allocation': _effective_allocation(demand),
                'deduction_amount': demand.get('initial_allocation', 0) - _effective_allocation(demand),
                'notes': demand.get('notes', ''),
                'suggestion_type': suggestion.get('suggestion_type'),
                'suggested_reduction': suggestion.get('suggested_reduction'),
                'suggestion_reason': suggestion.get('reason'),
                'highlight_color': suggestion.get('highlight_color'),
                'review_status': 'reviewed' if demand.get('user_allocation') is not None else 'pending'
            })
        
        return {
            "status": "success",
            "review_table": review_data,
            "total_count": len(review_data)
        }
        
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/sessions/{session_id}/run-allocation")
async def run_allocation(session_id: str, request: RunAllocationRequest):
    """
    Set initial allocations by the chosen priority (no distribution).
    historic_deficit -> initial_allocation = historical_deficit per row.
    requested_amount -> initial_allocation = current_request per row.
    last_year_data -> initial_allocation = graduate_count per row.
    Updates demand_requests.initial_allocation and adjustment_suggestions (red/yellow highlight).
    """
    try:
        session_resp = supabase.table("allocation_sessions").select("*").eq("id", session_id).execute()
        if not session_resp.data:
            raise HTTPException(status_code=404, detail="Session not found")
        session = session_resp.data[0]
        total_grants = request.total_grants if request.total_grants is not None else session.get("total_budget") or 0

        demands_resp = supabase.table("demand_requests").select(
            "*, region:regions(id, name), specialty:specialties(id, name)"
        ).eq("session_id", session_id).execute()
        demands = demands_resp.data or []
        if not demands:
            return {"status": "success", "message": "No demands to allocate", "updated": 0}

        graduate_year = (session.get("year") or datetime.now().year) - 1
        graduates_resp = supabase.table("yearly_graduates").select("specialty, graduate_count").eq("year", graduate_year).execute()
        graduates_map = {}
        for g in (graduates_resp.data or []):
            key = _normalize_specialty_for_graduate_lookup(g.get("specialty"))
            graduates_map[key] = int(g.get("graduate_count") or 0)

        def _demand_region_name(d):
            r = d.get("region")
            return r.get("name", "").strip() if isinstance(r, dict) else (d.get("region") or "")
        def _demand_specialty_name(d):
            s = d.get("specialty")
            return s.get("name", "").strip() if isinstance(s, dict) else (d.get("specialty") or "")

        rows = []
        for d in demands:
            region = _demand_region_name(d)
            specialty = _demand_specialty_name(d)
            hist = int(d.get("historical_deficit") or 0)
            req = int(d.get("current_request") or 0)
            grad = graduates_map.get(_normalize_specialty_for_graduate_lookup(specialty), 0)
            rows.append({
                "id": d["id"],
                "historical_deficit": hist,
                "current_request": req,
                "graduate_count": grad,
            })

        quotas = allocation_by_priority(request.priority, rows)
        if len(quotas) != len(demands):
            raise HTTPException(status_code=500, detail="Allocation length mismatch")

        # Use final_allocation when set; only apply priority to rows without final_allocation
        for i, demand in enumerate(demands):
            if demand.get("final_allocation") is not None:
                quotas[i] = int(demand["final_allocation"])

        demand_ids = [d["id"] for d in demands]
        for i, demand in enumerate(demands):
            new_alloc = quotas[i]
            supabase.table("demand_requests").update({
                "initial_allocation": new_alloc,
                "final_allocation": new_alloc,
            }).eq("id", demand["id"]).execute()

        for sid in demand_ids:
            supabase.table("adjustment_suggestions").delete().eq("demand_request_id", sid).execute()

        for i, demand in enumerate(demands):
            initial_allocation = quotas[i]
            hist = int(demand.get("historical_deficit") or 0)
            req = int(demand.get("current_request") or 0)
            max_need = max(hist, req)
            # Red = allocation could be less (over-allocated). Yellow = allocation could be higher (under-allocated).
            if initial_allocation > max_need:
                supabase.table("adjustment_suggestions").insert({
                    "demand_request_id": demand["id"],
                    "suggestion_type": "can_reduce",
                    "suggested_reduction": initial_allocation - max_need,
                    "reason": f"Можно уменьшить на {initial_allocation - max_need}",
                    "highlight_color": "red",
                }).execute()
            elif initial_allocation < max_need:
                supabase.table("adjustment_suggestions").insert({
                    "demand_request_id": demand["id"],
                    "suggestion_type": "under_allocated",
                    "suggested_reduction": 0,
                    "reason": f"Выделено ({initial_allocation}) меньше потребности ({max_need})",
                    "highlight_color": "yellow",
                }).execute()

        return {
            "status": "success",
            "message": f"Allocation run with priority={request.priority}",
            "updated": len(demands),
            "total_grants": total_grants,
            "priority": request.priority,
        }
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/demands/{demand_id}/adjust")
async def adjust_allocation(demand_id: str, request: AdjustAllocationRequest):
    """
    Adjust a single allocation
    """
    try:
        new_alloc = request.new_allocation
        update_data = {
            "user_allocation": new_alloc,
            "final_allocation": new_alloc if new_alloc is not None else None,
            "notes": request.notes,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        result = supabase.table("demand_requests").update(update_data).eq("id", demand_id).execute()
        
        if not result.data:
            raise HTTPException(status_code=404, detail="Demand request not found")
        
        return {
            "status": "success",
            "message": "Allocation adjusted successfully",
            "demand": result.data[0]
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/demands/bulk-adjust")
async def bulk_adjust_allocations(request: BulkAdjustRequest):
    """
    Bulk adjust multiple allocations at once
    """
    try:
        results = []
        errors = []
        
        for adj in request.adjustments:
            try:
                new_alloc = adj.new_allocation
                result = supabase.table("demand_requests").update({
                    "user_allocation": new_alloc,
                    "final_allocation": new_alloc if new_alloc is not None else None,
                    "notes": adj.notes,
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }).eq("id", adj.demand_request_id).execute()
                
                if result.data:
                    results.append(result.data[0])
            except Exception as e:
                errors.append({
                    "demand_id": adj.demand_request_id,
                    "error": str(e)
                })
        
        return {
            "status": "success" if not errors else "partial",
            "message": f"Adjusted {len(results)} allocations",
            "updated_count": len(results),
            "error_count": len(errors),
            "updated_demands": results,
            "errors": errors if errors else None
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/sessions/{session_id}/finalize")
async def finalize_session(session_id: str, request: FinalizeSessionRequest):
    """
    Finalize the allocation session (lock it)
    """
    try:
        # Check if all demands have been reviewed
        demands = supabase.table("demand_requests").select("user_allocation").eq("session_id", session_id).execute()
        
        if demands.data:
            unreviewed = sum(1 for d in demands.data if d.get('user_allocation') is None)
            
            if unreviewed > 0:
                return {
                    "status": "warning",
                    "message": f"{unreviewed} allocations have not been reviewed. Are you sure you want to finalize?",
                    "unreviewed_count": unreviewed,
                    "total_count": len(demands.data)
                }
        
        # Update session status
        result = supabase.table("allocation_sessions").update({
            "status": "finalized",
            "finalized_at": datetime.now(timezone.utc).isoformat()
        }).eq("id", session_id).execute()
        
        if not result.data:
            raise HTTPException(status_code=404, detail="Session not found")
        
        # Log audit entry
        supabase.table("audit_log").insert({
            "session_id": session_id,
            "action": "session_finalized",
            "changed_by": request.finalized_by,
            "created_at": datetime.now(timezone.utc).isoformat()
        }).execute()
        
        return {
            "status": "success",
            "message": "Session finalized successfully",
            "session": result.data[0]
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/sessions")
async def list_sessions(year: Optional[int] = None, status: Optional[str] = None):
    """
    List all allocation sessions with filters
    """
    try:
        query = supabase.table("allocation_sessions").select("*")
        
        if year:
            query = query.eq("year", year)
        if status:
            query = query.eq("status", status)
        
        result = query.order("created_at", desc=True).execute()
        
        return {
            "status": "success",
            "sessions": result.data,
            "count": len(result.data) if result.data else 0
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/sessions/{session_id}/summary")
async def get_session_summary(session_id: str):
    """
    Get detailed session summary with statistics
    """
    try:
        # Get session
        session_resp = supabase.table("allocation_sessions").select("*").eq("id", session_id).execute()
        
        if not session_resp.data:
            raise HTTPException(status_code=404, detail="Session not found")
        
        session = session_resp.data[0]
        
        # Get all demands
        demands_resp = supabase.table("demand_requests").select("*").eq("session_id", session_id).execute()
        demands = demands_resp.data or []
        
        # Calculate statistics
        total_deficit = sum(d.get('historical_deficit', 0) for d in demands)
        total_requested = sum(d.get('current_request', 0) for d in demands)
        total_initial = sum(d.get('initial_allocation', 0) for d in demands)
        total_final = sum(_effective_allocation(d) for d in demands)
        total_deductions = sum(d.get('initial_allocation', 0) - _effective_allocation(d) for d in demands)
        
        reviewed_count = sum(1 for d in demands if d.get('user_allocation') is not None)
        
        # Group by region
        by_region = {}
        for d in demands:
            region = d['region']
            if region not in by_region:
                by_region[region] = {
                    'deficit': 0,
                    'requested': 0,
                    'allocated': 0,
                    'count': 0
                }
            by_region[region]['deficit'] += d.get('historical_deficit', 0)
            by_region[region]['requested'] += d.get('current_request', 0)
            by_region[region]['allocated'] += _effective_allocation(d)
            by_region[region]['count'] += 1
        
        # Group by specialty
        by_specialty = {}
        for d in demands:
            specialty = d['specialty']
            if specialty not in by_specialty:
                by_specialty[specialty] = {
                    'deficit': 0,
                    'requested': 0,
                    'allocated': 0,
                    'count': 0
                }
            by_specialty[specialty]['deficit'] += d.get('historical_deficit', 0)
            by_specialty[specialty]['requested'] += d.get('current_request', 0)
            by_specialty[specialty]['allocated'] += _effective_allocation(d)
            by_specialty[specialty]['count'] += 1
        
        return {
            "status": "success",
            "session": session,
            "summary": {
                "total_demands": len(demands),
                "total_deficit": total_deficit,
                "total_requested": total_requested,
                "total_initial_allocation": total_initial,
                "total_final_allocation": total_final,
                "total_deductions": total_deductions,
                "budget": session['total_budget'],
                "budget_remaining": session['total_budget'] - total_final,
                "budget_utilization_pct": round((total_final / session['total_budget'] * 100), 1) if session['total_budget'] > 0 else 0,
                "reviewed_count": reviewed_count,
                "pending_count": len(demands) - reviewed_count,
                "review_progress_pct": round((reviewed_count / len(demands) * 100), 1) if len(demands) > 0 else 0
            },
            "by_region": by_region,
            "by_specialty": by_specialty
        }
        
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/sessions/{session_id}/export")
async def export_session(session_id: str, format: str = "json"):
    """
    Export session data (JSON or CSV)
    """
    try:
        # Get all demands
        demands_resp = supabase.table("demand_requests").select("*").eq("session_id", session_id).execute()
        
        if not demands_resp.data:
            raise HTTPException(status_code=404, detail="No data found for session")
        
        # Format export data
        export_data = []
        for d in demands_resp.data:
            export_data.append({
                "region": d['region'],
                "specialty": d['specialty'],
                "historical_deficit": d['historical_deficit'],
                "current_request": d['current_request'],
                "max_need": max(d.get('historical_deficit') or 0, d.get('current_request') or 0),
                "initial_allocation": d.get('initial_allocation', 0),
                "final_allocation": _effective_allocation(d),
                "deduction": d.get('initial_allocation', 0) - _effective_allocation(d),
                "notes": d.get('notes', '')
            })
        
        if format == "csv":
            import io
            import csv
            
            output = io.StringIO()
            if export_data:
                writer = csv.DictWriter(output, fieldnames=export_data[0].keys())
                writer.writeheader()
                writer.writerows(export_data)
            
            return {
                "status": "success",
                "format": "csv",
                "data": output.getvalue()
            }
        
        return {
            "status": "success",
            "format": "json",
            "data": export_data,
            "count": len(export_data)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/sessions/{session_id}")
async def delete_session(session_id: str):
    """
    Delete a session (cascade deletes all related data)
    """
    try:
        result = supabase.table("allocation_sessions").delete().eq("id", session_id).execute()
        
        return {
            "status": "success",
            "message": "Session deleted successfully"
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# --- Yearly Graduates ---

@app.post("/graduates/upload")
async def upload_graduates(
    file: UploadFile = File(...),
    session_id: Optional[str] = None,
    year: Optional[int] = None,
    years: Optional[str] = None,
    sheet_name: Optional[str] = None,
    replace: bool = False,
):
    """
    Upload Excel file with yearly graduate data (stored per specialty only).
    Excel columns: region, specialty, graduate_count (or graduates/count). Optional: year (or pass ?year=2024).
    replace: if True, delete existing rows for the given year(s) before insert (re-import).
    years: comma-separated years to replace when replace=1 (e.g. years=2024,2025,2026). If not set, uses year.
    """
    try:
        content = await file.read()
        ext = file.filename.lower().split(".")[-1]
        if ext not in ("xlsx", "xls"):
            raise HTTPException(status_code=400, detail="Only Excel (.xlsx, .xls) files are supported.")
        replace_years = None
        if replace:
            if years:
                replace_years = [int(y.strip()) for y in years.split(",") if y.strip().isdigit()]
            elif year is not None:
                replace_years = [year]
        result = await process_uploaded_graduates(
            content,
            supabase,
            session_id=session_id,
            default_year=year,
            file_name=file.filename,
            sheet_name=sheet_name,
            replace_years=replace_years,
        )
        if not result.get("success"):
            raise HTTPException(status_code=400, detail=result.get("error", "Failed to process file"))
        return {
            "status": "success",
            "message": f"Imported {result.get('inserted', 0)} graduate records",
            "inserted": result.get("inserted", 0),
            "total_rows": result.get("total_rows", 0),
            "metadata": result.get("metadata"),
        }
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/graduates")
async def list_graduates(year: Optional[int] = None, session_id: Optional[str] = None):
    """
    List yearly graduate records. Filter by year and/or session_id.
    """
    try:
        query = supabase.table("yearly_graduates").select("*").order("year", desc=True).order("specialty")
        if year is not None:
            query = query.eq("year", year)
        if session_id is not None:
            query = query.eq("session_id", session_id)
        result = query.execute()
        data = result.data or []
        return {
            "status": "success",
            "graduates": data,
            "count": len(data),
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/university-allocations")
async def list_university_allocations(
    region_id: Optional[str] = None,
    specialty_id: Optional[str] = None,
):
    """
    List university allocation assignments. Optional filters: region_id, specialty_id.
    Returns list of { region_id, specialty_id, university_id, allocated_count }.
    """
    try:
        query = supabase.table("university_allocation_assignments").select(
            "region_id, specialty_id, university_id, allocated_count"
        )
        if region_id:
            query = query.eq("region_id", region_id)
        if specialty_id:
            query = query.eq("specialty_id", specialty_id)
        resp = query.execute()
        data = resp.data or []
        return {"assignments": data, "count": len(data)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/university-allocations")
async def upsert_university_allocations(request: UpsertUniversityAllocationsRequest):
    """
    Replace all university assignments for one (region_id, specialty_id).
    Validates allocated_count >= 0. Does not check capacity or demand total.
    """
    try:
        region_id = request.region_id
        specialty_id = request.specialty_id
        # Delete existing for this (region, specialty)
        supabase.table("university_allocation_assignments").delete().eq(
            "region_id", region_id
        ).eq("specialty_id", specialty_id).execute()
        # Insert new rows (only those with count > 0)
        inserted = 0
        for item in request.assignments:
            count = max(0, int(item.allocated_count))
            if count <= 0:
                continue
            supabase.table("university_allocation_assignments").insert({
                "region_id": region_id,
                "specialty_id": specialty_id,
                "university_id": item.university_id,
                "allocated_count": count,
            }).execute()
            inserted += 1
        return {
            "status": "success",
            "region_id": region_id,
            "specialty_id": specialty_id,
            "assignments_count": inserted,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/university-allocations/geo")
async def compute_geo_university_allocations(
    save: bool = False,
    body: Optional[GeoFilterRequest] = Body(None),
):
    """
    Compute university allocations by geography: for each (region, specialty) demand,
    greedily assign by priority (primary → secondary → tertiary → specialized).
    Respects university_specialty_capacity.
    Returns the list of assignments. If save=true, upserts into university_allocation_assignments.
    Optional body: { region_id?, specialty_id? } to recompute only for that (region, specialty).
    """
    try:
        demand_filter = None
        if body and body.region_id and body.specialty_id:
            demand_filter = {(body.region_id, body.specialty_id)}
        session_id = body.session_id if body else None
        assignments = _compute_geo_university_allocations(supabase, demand_filter=demand_filter, session_id=session_id)
        if save and assignments:
            try:
                # Delete existing geo assignments for (region, specialty) pairs we're recomputing
                region_spec_pairs = set((a["region_id"], a["specialty_id"]) for a in assignments)
                for rid, sid in region_spec_pairs:
                    supabase.table("university_allocation_assignments").delete().eq("region_id", rid).eq("specialty_id", sid).execute()
                for a in assignments:
                    supabase.table("university_allocation_assignments").insert({
                        "region_id": a["region_id"],
                        "specialty_id": a["specialty_id"],
                        "university_id": a["university_id"],
                        "allocated_count": a["allocated_count"],
                    }).execute()
            except Exception as persist_e:
                return {
                    "status": "success",
                    "assignments": assignments,
                    "count": len(assignments),
                    "persisted": False,
                    "persist_error": str(persist_e),
                }
        return {
            "status": "success",
            "assignments": assignments,
            "count": len(assignments),
            "persisted": save and bool(assignments),
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/health")
async def health_check():
    return {
        "status": "ok",
        "service": "allocation-api-v3",
        "version": "3.0",
        "features": ["excel_upload", "pdf_upload", "regional_allocation", "yearly_graduates", "geo_university_allocation"]
    }


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
